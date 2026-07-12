from __future__ import annotations

import csv
import json
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Any


def _require_openpyxl():
    try:
        import openpyxl  # type: ignore
        return openpyxl
    except Exception as exc:  # pragma: no cover - depends on env
        raise RuntimeError('缺少 openpyxl，无法处理 XLSX 转换') from exc


def _require_yaml():
    try:
        import yaml  # type: ignore
        return yaml
    except Exception as exc:  # pragma: no cover
        raise RuntimeError('缺少 PyYAML，无法处理 YAML 转换') from exc


def _require_toml():
    try:
        import tomllib  # py3.11+
        return tomllib, None
    except Exception:
        try:
            import toml  # type: ignore
            return None, toml
        except Exception as exc:  # pragma: no cover
            raise RuntimeError('缺少 TOML 解析库，无法处理 TOML 转换') from exc


def read_delimited(path: Path, delimiter: str) -> list[dict[str, str]]:
    with path.open('r', encoding='utf-8-sig', newline='') as f:
        rows = list(csv.DictReader(f, delimiter=delimiter))
    return [dict(row) for row in rows]


def write_delimited(rows: list[dict[str, Any]], path: Path, delimiter: str) -> None:
    fields = sorted({key for row in rows for key in row.keys()})
    with path.open('w', encoding='utf-8', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=fields, delimiter=delimiter)
        writer.writeheader()
        writer.writerows(rows)


def read_json_rows(path: Path) -> list[dict[str, Any]]:
    data = json.loads(path.read_text(encoding='utf-8'))
    if isinstance(data, list):
        return [row if isinstance(row, dict) else {'value': row} for row in data]
    if isinstance(data, dict):
        for key in ('rows', 'data', 'items'):
            if isinstance(data.get(key), list):
                return [row if isinstance(row, dict) else {'value': row} for row in data[key]]
        return [data]
    return [{'value': data}]


def read_ndjson_rows(path: Path) -> list[dict[str, Any]]:
    rows = []
    for line in path.read_text(encoding='utf-8').splitlines():
        if not line.strip():
            continue
        item = json.loads(line)
        rows.append(item if isinstance(item, dict) else {'value': item})
    return rows


def read_xlsx_rows(path: Path) -> list[dict[str, Any]]:
    openpyxl = _require_openpyxl()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(cell or f'column_{index + 1}') for index, cell in enumerate(rows[0])]
    result = []
    for row in rows[1:]:
        result.append({headers[index]: value for index, value in enumerate(row) if index < len(headers)})
    return result


def write_xlsx(rows: list[dict[str, Any]], path: Path) -> None:
    openpyxl = _require_openpyxl()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    fields = sorted({key for row in rows for key in row.keys()}) or ['value']
    ws.append(fields)
    for row in rows:
        ws.append([row.get(field) for field in fields])
    wb.save(path)


def xml_to_dict(element: ET.Element) -> dict[str, Any]:
    children = list(element)
    if not children:
        return {element.tag: element.text or ''}
    result: dict[str, Any] = {}
    for child in children:
        child_value = xml_to_dict(child)
        key, value = next(iter(child_value.items()))
        if key in result:
            if not isinstance(result[key], list):
                result[key] = [result[key]]
            result[key].append(value)
        else:
            result[key] = value
    return {element.tag: result}


def dict_to_xml(name: str, value: Any) -> ET.Element:
    element = ET.Element(name)
    if isinstance(value, dict):
        for key, child_value in value.items():
            element.append(dict_to_xml(str(key), child_value))
    elif isinstance(value, list):
        for item in value:
            element.append(dict_to_xml('item', item))
    else:
        element.text = '' if value is None else str(value)
    return element


def read_rows(source: str, path: Path) -> list[dict[str, Any]]:
    if source == 'txt':
        text = path.read_text(encoding='utf-8-sig')
        first_line = text.splitlines()[0] if text.splitlines() else ''
        delimiter = '\t' if '\t' in first_line else ','
        rows = []
        for index, line in enumerate(text.splitlines(), start=1):
            if not line.strip():
                continue
            parts = line.split(delimiter) if delimiter in line else [line]
            rows.append({f'column_{i + 1}': value for i, value in enumerate(parts)})
        return rows
    if source == 'csv':
        return read_delimited(path, ',')
    if source == 'tsv':
        return read_delimited(path, '\t')
    if source == 'json':
        return read_json_rows(path)
    if source == 'ndjson':
        return read_ndjson_rows(path)
    if source == 'xlsx':
        return read_xlsx_rows(path)
    raise RuntimeError(f'数据读取暂不支持 {source}')


def convert_data(source: str, target: str, input_path: Path, output_path: Path) -> list[str]:
    logs = [f'读取 {source.upper()} 数据']

    if source == 'json' and target == 'yaml':
        yaml = _require_yaml()
        data = json.loads(input_path.read_text(encoding='utf-8'))
        output_path.write_text(yaml.safe_dump(data, allow_unicode=True, sort_keys=False), encoding='utf-8')
        return logs + ['JSON 已转换为 YAML']

    if source == 'json' and target == 'toml':
        _, toml = _require_toml()
        if not toml:
            raise RuntimeError('当前环境只能读取 TOML，写出 TOML 需要安装 toml 包')
        data = json.loads(input_path.read_text(encoding='utf-8'))
        output_path.write_text(toml.dumps(data), encoding='utf-8')
        return logs + ['JSON 已转换为 TOML']

    if source == 'json' and target == 'xml':
        data = json.loads(input_path.read_text(encoding='utf-8'))
        root = dict_to_xml('root', data)
        ET.ElementTree(root).write(output_path, encoding='utf-8', xml_declaration=True)
        return logs + ['JSON 已转换为 XML']





    if source in {'csv', 'tsv', 'json', 'ndjson', 'xlsx', 'txt'}:
        rows = read_rows(source, input_path)
        logs.append(f'读取到 {len(rows)} 行')
        if target == 'csv':
            write_delimited(rows, output_path, ',')
        elif target == 'tsv':
            write_delimited(rows, output_path, '\t')
        elif target == 'txt':
            write_delimited(rows, output_path, '\t')
        elif target == 'json':
            output_path.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding='utf-8')
        elif target == 'xlsx':
            write_xlsx(rows, output_path)
        else:
            raise RuntimeError(f'数据写出暂不支持 {target}')
        logs.append(f'已写出 {target.upper()} 文件')
        return logs

    if source == 'yaml' and target == 'json':
        yaml = _require_yaml()
        data = yaml.safe_load(input_path.read_text(encoding='utf-8'))
        output_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding='utf-8')
        return logs + ['YAML 已转换为 JSON']


    if source == 'toml' and target == 'json':
        tomllib, toml = _require_toml()
        if tomllib:
            data = tomllib.loads(input_path.read_text(encoding='utf-8'))
        else:
            data = toml.loads(input_path.read_text(encoding='utf-8'))
        output_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding='utf-8')
        return logs + ['TOML 已转换为 JSON']


    if source == 'xml' and target == 'json':
        root = ET.parse(input_path).getroot()
        output_path.write_text(json.dumps(xml_to_dict(root), ensure_ascii=False, indent=2), encoding='utf-8')
        return logs + ['XML 已转换为 JSON']


    raise RuntimeError(f'暂不支持 {source} → {target}')
