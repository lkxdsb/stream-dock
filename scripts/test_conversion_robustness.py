#!/usr/bin/env python3
"""生成多类型样例文件，并对 StreamDock 本地文件转换做鲁棒性/准确性验收。

默认在 /tmp 下创建隔离工作区，生成 fixture、执行转换、做内容级断言，并写出 JSON 报告。
用法：
  python scripts/test_conversion_robustness.py
  python scripts/test_conversion_robustness.py --keep --workdir /tmp/streamdock-convert-lab
"""
from __future__ import annotations

import argparse
import csv
import gzip
import json
import os
import shutil
import subprocess
import sys
import tarfile
import tempfile
import textwrap
import time
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable, Any

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from converters.pipeline import convert_file  # noqa: E402
from runtime_checks import augmented_path, resolve_tool_path, validate_media_output  # noqa: E402

os.environ['PATH'] = augmented_path()


@dataclass
class Case:
    name: str
    source: str
    target: str
    input_path: Path
    validate: Callable[[Path], None]
    expected_error: str | None = None
    notes: str = ''


@dataclass
class CaseResult:
    name: str
    source: str
    target: str
    status: str
    elapsed_seconds: float
    output_path: str | None = None
    error: str | None = None
    validation: dict[str, Any] = field(default_factory=dict)
    notes: str = ''


def require(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding='utf-8'))


def generate_fixtures(root: Path) -> dict[str, Path]:
    fixtures = root / 'fixtures'
    fixtures.mkdir(parents=True, exist_ok=True)
    paths: dict[str, Path] = {}

    rows = [
        {'id': '1', 'name': '张三', 'score': '98.5', 'remark': '含逗号, 引号"与 emoji ✅'},
        {'id': '2', 'name': '李四', 'score': '87', 'remark': '第二行中文内容'},
        {'id': '3', 'name': 'Ada', 'score': '100', 'remark': 'ASCII fallback'},
    ]
    paths['csv'] = fixtures / 'table_utf8.csv'
    with paths['csv'].open('w', encoding='utf-8-sig', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=['id', 'name', 'score', 'remark'])
        writer.writeheader(); writer.writerows(rows)

    paths['tsv'] = fixtures / 'table_utf8.tsv'
    with paths['tsv'].open('w', encoding='utf-8', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=['id', 'name', 'score', 'remark'], delimiter='\t')
        writer.writeheader(); writer.writerows(rows)

    paths['json'] = fixtures / 'rows.json'
    paths['json'].write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding='utf-8')

    paths['ndjson'] = fixtures / 'rows.ndjson'
    paths['ndjson'].write_text('\n'.join(json.dumps(row, ensure_ascii=False) for row in rows) + '\n', encoding='utf-8')

    paths['yaml'] = fixtures / 'config.yaml'
    paths['yaml'].write_text('title: 鲁棒性测试\nitems:\n  - name: 张三\n    ok: true\n', encoding='utf-8')

    paths['toml'] = fixtures / 'config.toml'
    paths['toml'].write_text('[project]\nname = "StreamDock"\nowner = "张三"\ncount = 3\n', encoding='utf-8')

    paths['xml'] = fixtures / 'items.xml'
    paths['xml'].write_text('<?xml version="1.0" encoding="utf-8"?><root><item><name>张三</name><score>98</score></item></root>', encoding='utf-8')

    try:
        import openpyxl  # type: ignore
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = '数据'
        ws.append(['id', 'name', 'score', 'remark'])
        for row in rows:
            ws.append([row['id'], row['name'], float(row['score']), row['remark']])
        paths['xlsx'] = fixtures / 'table.xlsx'
        wb.save(paths['xlsx'])
    except Exception:
        pass

    paths['md'] = fixtures / 'readme.md'
    paths['md'].write_text(textwrap.dedent('''\
        # StreamDock 转换测试

        这是一段中文 Markdown，包含 **加粗**、列表和表格。

        | 姓名 | 分数 |
        | --- | ---: |
        | 张三 | 98 |
        | Ada | 100 |
        '''), encoding='utf-8')

    paths['html'] = fixtures / 'page.html'
    paths['html'].write_text('<!doctype html><meta charset="utf-8"><h1>标题</h1><p>中文段落 ✅</p>', encoding='utf-8')

    paths['txt'] = fixtures / 'plain.txt'
    paths['txt'].write_text('第一行中文\nSecond line ✅\n第三行\n', encoding='utf-8')

    paths['rtf'] = fixtures / 'sample.rtf'
    paths['rtf'].write_text(r'{\rtf1\ansi StreamDock \par plain text paragraph}', encoding='utf-8')

    paths['srt'] = fixtures / 'caption.srt'
    paths['srt'].write_text('1\n00:00:00,000 --> 00:00:01,500\n你好 StreamDock\n\n2\n00:00:01,500 --> 00:00:03,000\n第二句字幕\n', encoding='utf-8')
    paths['vtt'] = fixtures / 'caption.vtt'
    paths['vtt'].write_text('WEBVTT\n\n00:00:00.000 --> 00:00:01.000\n你好 VTT\n', encoding='utf-8')
    paths['lrc'] = fixtures / 'lyric.lrc'
    paths['lrc'].write_text('[00:00.00]第一句歌词\n[00:02.50]第二句歌词\n', encoding='utf-8')
    paths['ass'] = fixtures / 'caption.ass'
    paths['ass'].write_text('[Events]\nDialogue: 0,0:00:00.00,0:00:01.20,Default,,0,0,0,,你好\\NASS\n', encoding='utf-8')

    try:
        from PIL import Image, ImageDraw  # type: ignore
        img = Image.new('RGBA', (160, 90), (30, 80, 160, 255))
        draw = ImageDraw.Draw(img)
        draw.rectangle((10, 10, 150, 80), fill=(255, 220, 60, 210))
        draw.ellipse((55, 20, 105, 70), fill=(230, 50, 80, 230))
        paths['png'] = fixtures / 'cover.png'; img.save(paths['png'])
        img.convert('RGB').save(fixtures / 'cover.jpg'); paths['jpg'] = fixtures / 'cover.jpg'
        frames = []
        for idx, color in enumerate(((255, 0, 0), (0, 180, 80), (60, 80, 255))):
            frame = Image.new('RGB', (64, 64), color)
            ImageDraw.Draw(frame).text((8, 24), str(idx + 1), fill=(255, 255, 255))
            frames.append(frame)
        paths['gif'] = fixtures / 'anim.gif'
        frames[0].save(paths['gif'], save_all=True, append_images=frames[1:], duration=80, loop=0)
    except Exception:
        pass

    archive_src = fixtures / 'folder_src'
    (archive_src / 'nested').mkdir(parents=True, exist_ok=True)
    (archive_src / 'nested' / '中文.txt').write_text('archive ok ✅', encoding='utf-8')
    (archive_src / 'numbers.csv').write_text('n\n1\n2\n', encoding='utf-8')
    paths['folder'] = archive_src
    paths['zip'] = fixtures / 'sample.zip'
    with zipfile.ZipFile(paths['zip'], 'w', zipfile.ZIP_DEFLATED) as zf:
        for p in archive_src.rglob('*'):
            if p.is_file():
                zf.write(p, p.relative_to(archive_src))
    paths['tar.gz'] = fixtures / 'sample.tar.gz'
    with tarfile.open(paths['tar.gz'], 'w:gz') as tf:
        tf.add(archive_src, arcname='folder_src')
    paths['gz'] = fixtures / 'plain.txt.gz'
    with gzip.open(paths['gz'], 'wb') as f:
        f.write(paths['txt'].read_bytes())
    paths['unsafe_zip'] = fixtures / 'unsafe.zip'
    with zipfile.ZipFile(paths['unsafe_zip'], 'w') as zf:
        zf.writestr('../escape.txt', 'blocked')

    ffmpeg = resolve_tool_path('ffmpeg')
    if shutil.which('ffmpeg', path=augmented_path()):
        paths['wav'] = fixtures / 'tone.wav'
        subprocess.run([ffmpeg, '-y', '-hide_banner', '-loglevel', 'error', '-f', 'lavfi', '-i', 'sine=frequency=880:duration=0.7', str(paths['wav'])], check=True, env={**os.environ, 'PATH': augmented_path()})
        paths['mp4'] = fixtures / 'sample.mp4'
        subprocess.run([
            ffmpeg, '-y', '-hide_banner', '-loglevel', 'error',
            '-f', 'lavfi', '-i', 'testsrc=size=160x90:rate=10:duration=1',
            '-f', 'lavfi', '-i', 'sine=frequency=440:duration=1',
            '-c:v', 'libx264', '-pix_fmt', 'yuv420p', '-c:a', 'aac', '-shortest', str(paths['mp4'])
        ], check=True, env={**os.environ, 'PATH': augmented_path()})

    return paths


def build_cases(paths: dict[str, Path]) -> list[Case]:
    cases: list[Case] = []

    def has(*keys: str) -> bool:
        return all(key in paths for key in keys)

    if has('csv'):
        cases.append(Case('CSV 保留中文/emoji到 JSON', 'csv', 'json', paths['csv'], lambda p: require(any(row.get('name') == '张三' and 'emoji' in row.get('remark', '') for row in read_json(p)), 'CSV→JSON 内容丢失')))
        cases.append(Case('CSV 到 XLSX 表格行数', 'csv', 'xlsx', paths['csv'], validate_xlsx_contains_zhangsan))
    if has('xlsx'):
        cases.append(Case('XLSX 到 CSV 内容回读', 'xlsx', 'csv', paths['xlsx'], validate_csv_contains_zhangsan))
    if has('json'):
        cases.append(Case('JSON 到 CSV 字段展开', 'json', 'csv', paths['json'], validate_csv_contains_zhangsan))
        cases.append(Case('JSON 到 YAML Unicode', 'json', 'yaml', paths['json'], lambda p: require('张三' in p.read_text(encoding='utf-8'), 'JSON→YAML 中文丢失')))
    if has('ndjson'):
        cases.append(Case('NDJSON 到 CSV 多行', 'ndjson', 'csv', paths['ndjson'], validate_csv_contains_zhangsan))
    if has('yaml'):
        cases.append(Case('YAML 到 JSON 嵌套结构', 'yaml', 'json', paths['yaml'], lambda p: require(read_json(p)['items'][0]['name'] == '张三', 'YAML→JSON 嵌套值错误')))
    if has('toml'):
        cases.append(Case('TOML 到 JSON 嵌套结构', 'toml', 'json', paths['toml'], lambda p: require(read_json(p)['project']['owner'] == '张三', 'TOML→JSON 值错误')))
    if has('xml'):
        cases.append(Case('XML 到 JSON 标签结构', 'xml', 'json', paths['xml'], lambda p: require(read_json(p)['root']['item']['name'] == '张三', 'XML→JSON 标签值错误')))

    if has('md'):
        cases.append(Case('Markdown 到 HTML 表格/中文', 'md', 'html', paths['md'], lambda p: require('<table>' in p.read_text(encoding='utf-8') and '张三' in p.read_text(encoding='utf-8'), 'MD→HTML 表格或中文缺失')))
        cases.append(Case('Markdown 到 PDF 非空', 'md', 'pdf', paths['md'], lambda p: require(p.stat().st_size > 1000, 'MD→PDF 文件过小')))
        cases.append(Case('Markdown 到 DOCX 段落', 'md', 'docx', paths['md'], validate_docx_contains_streamdock))
    if has('html'):
        cases.append(Case('HTML 到 TXT 去标签', 'html', 'txt', paths['html'], lambda p: require('中文段落' in p.read_text(encoding='utf-8') and '<p>' not in p.read_text(encoding='utf-8'), 'HTML→TXT 去标签失败')))
    if has('txt'):
        cases.append(Case('TXT 到 RTF 基础文本', 'txt', 'rtf', paths['txt'], lambda p: require(r'\rtf1' in p.read_text(encoding='utf-8'), 'TXT→RTF 头缺失')))
    if has('rtf'):
        cases.append(Case('RTF 到 TXT 文本抽取', 'rtf', 'txt', paths['rtf'], lambda p: require('StreamDock' in p.read_text(encoding='utf-8'), 'RTF→TXT 文本缺失')))

    if has('png'):
        cases.append(Case('PNG 到 JPG 尺寸/模式', 'png', 'jpg', paths['png'], lambda p: validate_image(p, (160, 90))))
        cases.append(Case('PNG 到 WEBP 尺寸/格式', 'png', 'webp', paths['png'], lambda p: validate_image(p, (160, 90))))
        cases.append(Case('PNG 到 ICO 非空', 'png', 'ico', paths['png'], lambda p: require(p.stat().st_size > 100, 'PNG→ICO 文件过小')))
    if has('gif'):
        cases.append(Case('GIF 到 PNG 帧目录', 'gif', 'png', paths['gif'], validate_gif_frames))

    if has('srt'):
        cases.append(Case('SRT 到 VTT 时间格式', 'srt', 'vtt', paths['srt'], lambda p: require(p.read_text(encoding='utf-8').startswith('WEBVTT') and '00:00:00.000' in p.read_text(encoding='utf-8'), 'SRT→VTT 格式错误')))
    if has('vtt'):
        cases.append(Case('VTT 到 SRT 时间格式', 'vtt', 'srt', paths['vtt'], lambda p: require('00:00:00,000' in p.read_text(encoding='utf-8'), 'VTT→SRT 时间格式错误')))
    if has('lrc'):
        cases.append(Case('LRC 到 SRT 歌词时间轴', 'lrc', 'srt', paths['lrc'], lambda p: require('第一句歌词' in p.read_text(encoding='utf-8') and '00:00:02,500' in p.read_text(encoding='utf-8'), 'LRC→SRT 内容错误')))
    if has('ass'):
        cases.append(Case('ASS 到 SRT 换行降级', 'ass', 'srt', paths['ass'], lambda p: require('你好\nASS' in p.read_text(encoding='utf-8'), 'ASS→SRT 换行/文本错误')))

    if has('folder'):
        cases.append(Case('文件夹到 ZIP 目录结构', 'folder', 'zip', paths['folder'], validate_zip_contains_nested))
        cases.append(Case('文件夹到 TAR.GZ 目录结构', 'folder', 'tar.gz', paths['folder'], lambda p: require(tarfile.is_tarfile(p), 'folder→tar.gz 不是有效 tar')))
    if has('zip'):
        cases.append(Case('ZIP 到文件夹安全解压', 'zip', 'folder', paths['zip'], lambda p: require((p / 'nested' / '中文.txt').exists(), 'ZIP→folder 文件缺失')))
    if has('tar.gz'):
        cases.append(Case('TAR.GZ 到 ZIP 重打包', 'tar.gz', 'zip', paths['tar.gz'], lambda p: require(zipfile.is_zipfile(p), 'tar.gz→zip 不是有效 zip')))
    if has('gz'):
        cases.append(Case('GZ 单文件解压', 'gz', 'folder', paths['gz'], lambda p: require((p / 'plain.txt').exists(), 'GZ→folder 解压文件缺失')))
    if has('unsafe_zip'):
        cases.append(Case('恶意 ZIP 路径穿越拒绝', 'zip', 'folder', paths['unsafe_zip'], lambda p: None, expected_error='不安全路径'))

    if has('wav'):
        cases.append(Case('WAV 到 MP3 音频流', 'wav', 'mp3', paths['wav'], lambda p: require(validate_media_output(p, expected_kind='audio')['hasAudio'], 'WAV→MP3 未检测到音频')))
    if has('mp4'):
        cases.append(Case('MP4 到 MP3 音频提取', 'mp4', 'mp3', paths['mp4'], lambda p: require(validate_media_output(p, expected_kind='audio')['hasAudio'], 'MP4→MP3 未检测到音频')))
        cases.append(Case('MP4 到 GIF 动图输出', 'mp4', 'gif', paths['mp4'], lambda p: require(p.stat().st_size > 1000 and p.read_bytes()[:3] == b'GIF', 'MP4→GIF 输出异常')))
        cases.append(Case('MP4 到 WEBM 视频流', 'mp4', 'webm', paths['mp4'], lambda p: require(validate_media_output(p, expected_kind='video')['hasVideo'], 'MP4→WEBM 未检测到视频')))

    return cases


def validate_csv_contains_zhangsan(path: Path) -> None:
    text = path.read_text(encoding='utf-8')
    require('张三' in text and 'remark' in text, 'CSV 内容缺少中文或表头')


def validate_xlsx_contains_zhangsan(path: Path) -> None:
    import openpyxl  # type: ignore
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    values = [cell for row in wb.active.iter_rows(values_only=True) for cell in row]
    require('张三' in values and len(values) >= 8, 'XLSX 内容或行列数异常')


def validate_docx_contains_streamdock(path: Path) -> None:
    from docx import Document  # type: ignore
    text = '\n'.join(p.text for p in Document(str(path)).paragraphs)
    require('StreamDock 转换测试' in text and '张三' in text, 'DOCX 段落内容缺失')


def validate_image(path: Path, size: tuple[int, int]) -> None:
    from PIL import Image  # type: ignore
    with Image.open(path) as img:
        require(img.size == size, f'图片尺寸异常：{img.size} != {size}')


def validate_gif_frames(path: Path) -> None:
    require(path.is_dir(), 'GIF→PNG 应输出帧目录')
    frames = sorted(path.glob('frame_*.png'))
    require(len(frames) >= 3, f'GIF 帧数不足：{len(frames)}')
    validate_image(frames[0], (64, 64))


def validate_zip_contains_nested(path: Path) -> None:
    require(zipfile.is_zipfile(path), '不是有效 ZIP')
    with zipfile.ZipFile(path) as zf:
        names = set(zf.namelist())
    require('nested/中文.txt' in names and 'numbers.csv' in names, 'ZIP 目录结构不完整')


def run_case(case: Case, output_dir: Path) -> CaseResult:
    start = time.perf_counter()
    result = convert_file(case.input_path, case.input_path.name, case.source, case.target, output_dir, naming_strategy='append')
    elapsed = round(time.perf_counter() - start, 3)
    if case.expected_error:
        if not result.success and result.error and case.expected_error in result.error:
            return CaseResult(case.name, case.source, case.target, 'PASS', elapsed, error=result.error, notes='按预期拒绝异常输入')
        return CaseResult(case.name, case.source, case.target, 'FAIL', elapsed, output_path=str(result.output_path) if result.output_path else None, error=f'期望错误包含 {case.expected_error!r}，实际 success={result.success}, error={result.error!r}')
    if not result.success or not result.output_path:
        return CaseResult(case.name, case.source, case.target, 'FAIL', elapsed, error=result.error or '未生成输出')
    try:
        case.validate(result.output_path)
    except Exception as exc:
        return CaseResult(case.name, case.source, case.target, 'FAIL', elapsed, output_path=str(result.output_path), error=str(exc), validation=result.validation or {}, notes=case.notes)
    return CaseResult(case.name, case.source, case.target, 'PASS', elapsed, output_path=str(result.output_path), validation=result.validation or {}, notes=case.notes)


def main() -> int:
    parser = argparse.ArgumentParser(description='StreamDock 文件转换鲁棒性/准确性验证脚本')
    parser.add_argument('--workdir', type=Path, default=None, help='工作目录，默认创建临时目录')
    parser.add_argument('--keep', action='store_true', help='保留工作目录，便于排查')
    parser.add_argument('--report', type=Path, default=ROOT / 'report_figures' / 'conversion_robustness_latest.json', help='JSON 报告输出路径')
    args = parser.parse_args()

    temp_ctx = None
    if args.workdir:
        workdir = args.workdir.expanduser().resolve()
        workdir.mkdir(parents=True, exist_ok=True)
    else:
        temp_ctx = tempfile.TemporaryDirectory(prefix='streamdock-conversion-robustness-')
        workdir = Path(temp_ctx.name)
    output_dir = workdir / 'outputs'; output_dir.mkdir(parents=True, exist_ok=True)

    try:
        paths = generate_fixtures(workdir)
        cases = build_cases(paths)
        results = [run_case(case, output_dir) for case in cases]
        passed = sum(1 for item in results if item.status == 'PASS')
        failed = [item for item in results if item.status != 'PASS']
        report = {
            'generatedAt': time.strftime('%Y-%m-%d %H:%M:%S'),
            'workdir': str(workdir),
            'total': len(results),
            'passed': passed,
            'failed': len(failed),
            'results': [item.__dict__ for item in results],
        }
        args.report.parent.mkdir(parents=True, exist_ok=True)
        args.report.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding='utf-8')

        print(f'工作目录: {workdir}')
        print(f'报告文件: {args.report}')
        print(f'转换用例: {passed}/{len(results)} 通过')
        for item in results:
            symbol = '✓' if item.status == 'PASS' else '✗'
            print(f'{symbol} {item.name} [{item.source}->{item.target}] {item.elapsed_seconds:.3f}s')
            if item.error and item.status != 'PASS':
                print(f'  error: {item.error}')
        if failed:
            print('\n失败用例请查看 JSON 报告中的 output_path/error。')
            return 1
        return 0
    finally:
        if temp_ctx and not args.keep:
            temp_ctx.cleanup()


if __name__ == '__main__':
    raise SystemExit(main())
